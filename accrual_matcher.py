"""
Accrual-to-Reversal Matching Script (Incremental Mode)
=======================================================
Reads a previously-matched workbook with new transaction rows pasted at the
bottom, matches the new rows against any unmatched existing entries, then
writes a timestamped output file with all match IDs preserved.

WORKFLOW:
    1. Open your existing matched workbook in Excel
    2. Paste new GL rows at the bottom of the data sheet, leaving the
       MatchGroup, MatchStatus, MatchedWith, NetCheck, and RowType columns
       BLANK on the new rows
    3. Save the file
    4. Run this script against it

The script will:
    - Preserve every existing MatchGroup ID, status, and net check
    - Classify and match only new rows + previously-unmatched rows
    - Continue match IDs from where the prior run left off
       (e.g. AR-0649 -> AR-0650, AI-0650 -> AI-0651)
    - Try to pair new reversals against still-open accruals from prior periods
    - Save to a new file with a timestamp suffix

USAGE:
    python accrual_matcher.py <input_file> [sheet_name] [vendor_csv]

EXAMPLES:
    python accrual_matcher.py Transactions_to_Clean.xlsx
    python accrual_matcher.py Transactions_to_Clean.xlsx "Accrual - Data (Matched Detail)"
    python accrual_matcher.py Transactions_to_Clean.xlsx "" "Legal List.csv"

The vendor CSV is expected to have at least a "Vendor" column.  If a
"Category" column is present, matched rows will also get that category.
Default path (if not supplied):
    C:\\Users\\chammer\\OneDrive - Red Roof\\Desktop\\Accrual Matcher\\Legal List.csv

If no sheet name is provided, the script tries these defaults in order:
    1. "Accrual - Data (Matched Detail)"
    2. "Matched Detail"
    3. "Clean Detail"
    4. The first sheet in the workbook

OUTPUT:
    A new file named <input_basename>_<YYYYMMDD_HHMMSS>.xlsx in the same
    directory, with all formatting and match IDs preserved.

Requirements:
    pip install pandas openpyxl
"""

import sys
import re
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------
DEFAULT_SHEET_NAMES = [
    "Accrual - Data (Matched Detail)",
    "Matched Detail",
    "Clean Detail",
]

MATCHING_COLS = ["MatchGroup", "MatchStatus", "MatchedWith", "NetCheck", "RowType"]

DEFAULT_VENDOR_CSV = (
    r"C:\Users\chammer\OneDrive - Red Roof\Desktop\Accrual Matcher\Legal List.csv"
)

# Column name mapping: maps the names used internally by the script to
# possible header names found in source files.  The script will auto-detect
# which variant is present and rename on load.
COLUMN_ALIASES = {
    "Mapped Vendor": ["Vendor Name"],
    "Amount": ["TOTAL"],
}


# ---------------------------------------------------------------------------
# 0. NORMALISE INCOMING COLUMNS
# ---------------------------------------------------------------------------
def normalise_columns(df):
    """
    Ensure the DataFrame has the internal column names the matching engine
    expects.  Where a column is missing but an alias exists, rename it.
    Where a column must be derived, compute it.
    """
    # Apply simple renames from COLUMN_ALIASES
    rename_map = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        if canonical not in df.columns:
            for alias in aliases:
                if alias in df.columns:
                    rename_map[alias] = canonical
                    break
    if rename_map:
        df = df.rename(columns=rename_map)

    # Derive Amount from ENTERED_DR / ENTERED_CR if still missing
    if "Amount" not in df.columns:
        if "ENTERED_DR" in df.columns and "ENTERED_CR" in df.columns:
            df["ENTERED_DR"] = pd.to_numeric(df["ENTERED_DR"], errors="coerce").fillna(0)
            df["ENTERED_CR"] = pd.to_numeric(df["ENTERED_CR"], errors="coerce").fillna(0)
            df["Amount"] = df["ENTERED_DR"] - df["ENTERED_CR"]
        else:
            df["Amount"] = 0

    # If Company column is absent, fill with a constant so matching logic
    # still groups correctly (all rows treated as same company).
    if "Company" not in df.columns:
        df["Company"] = "ALL"

    return df


# ---------------------------------------------------------------------------
# 0b. VENDOR EXTRACTION FROM GL DESCRIPTION
# ---------------------------------------------------------------------------
def load_vendor_list(csv_path):
    """
    Load the vendor reference CSV.  Returns a list of (vendor, category) tuples
    sorted longest-vendor-first so partial matching prefers the most specific hit.
    """
    if not os.path.isfile(csv_path):
        print(f"Warning: vendor list not found at {csv_path} — skipping vendor extraction")
        return []

    vdf = pd.read_csv(csv_path, encoding="cp1252")
    if "Vendor" not in vdf.columns:
        print(f"Warning: vendor CSV has no 'Vendor' column — skipping vendor extraction")
        return []

    has_category = "Category" in vdf.columns
    entries = []
    for _, row in vdf.iterrows():
        vendor = str(row["Vendor"]).strip()
        if not vendor or vendor.lower() in ("nan", ""):
            continue
        category = str(row["Category"]).strip() if has_category else ""
        entries.append((vendor, category))

    # Sort longest first so "WILSON, ELSER, MOSKOWITZ, EDELMAN & DICKER, LLP"
    # matches before a hypothetical shorter substring like "WILSON"
    entries.sort(key=lambda x: len(x[0]), reverse=True)
    return entries


def extract_vendor_from_description(description, vendor_list):
    """
    Case-insensitive partial match of known vendor names against a GL Description.
    Returns (vendor, category) on first (longest) hit, or (None, None).
    """
    desc_upper = str(description).upper()
    for vendor, category in vendor_list:
        if vendor.upper() in desc_upper:
            return vendor, category
    return None, None


def apply_vendor_extraction(df, vendor_list):
    """
    Scan GL Description for known vendor names.  Always overwrites Mapped Vendor
    when a match is found.  Also sets Category if the vendor list provides one.
    """
    if not vendor_list:
        return df

    if "GL Description" not in df.columns:
        print("  No 'GL Description' column found — skipping vendor extraction")
        return df

    if "Mapped Vendor" not in df.columns:
        df["Mapped Vendor"] = ""
    if "Category" not in df.columns:
        df["Category"] = ""

    matched_count = 0
    for idx, row in df.iterrows():
        desc = row.get("GL Description", "")
        if is_blank(desc):
            continue
        vendor, category = extract_vendor_from_description(desc, vendor_list)
        if vendor is not None:
            df.loc[idx, "Mapped Vendor"] = vendor
            if category:
                df.loc[idx, "Category"] = category
            matched_count += 1

    print(f"  Vendor extraction: {matched_count} rows matched from GL Description")
    return df


# ---------------------------------------------------------------------------
# 1. ROW CLASSIFICATION
# ---------------------------------------------------------------------------
def classify_row(row):
    """Assign a type to each GL row based on GL Journal text and GL Source."""
    journal = str(row["GL Journal"])
    source = str(row.get("GL Source", ""))

    if journal.startswith("Reverses "):
        return "Reversal"
    if "Accrual" in journal and source == "Spreadsheet":
        return "Accrual"
    if source == "Payables":
        return "Invoice"
    if "Reclass" in journal:
        return "Reclass"
    if "Adjustment" in journal:
        return "Adjustment"
    if "Other" in journal:
        return "Other"
    if "Amortization" in journal:
        return "Amortization"
    return "Other"


# ---------------------------------------------------------------------------
# 2. EXTRACT ACCRUAL REFERENCE FROM A REVERSAL'S GL JOURNAL
# ---------------------------------------------------------------------------
def extract_accrual_ref(journal):
    """
    Parse the original accrual journal name out of a reversal entry.
    Example input : "Reverses ALA APY 1224 AP ACCRUALS 010325 Accrual 27-01-25 16:08:53"
    Example output: "ALA APY 1224 AP ACCRUALS 010325 Accrual"
    """
    m = re.match(
        r"Reverses (.+?)\s+\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}",
        str(journal),
    )
    return m.group(1).strip() if m else None


# ---------------------------------------------------------------------------
# 3. IDENTIFY WHICH ROWS NEED MATCHING
# ---------------------------------------------------------------------------
def is_blank(value):
    """Treat NaN, None, and empty strings as blank."""
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def needs_matching(row):
    """
    A row needs matching if it has no MatchGroup.
    This catches both freshly-pasted rows and previously-unmatched rows.
    """
    has_group = not is_blank(row.get("MatchGroup"))
    return not has_group


# ---------------------------------------------------------------------------
# 4. PARSE EXISTING MATCH ID COUNTERS
# ---------------------------------------------------------------------------
def get_next_match_ids(df):
    """Find the highest existing AR-#### and AI-#### so new IDs continue cleanly."""
    ar_max = 0
    ai_max = 0
    if "MatchGroup" not in df.columns:
        return 1, 1
    for tag in df["MatchGroup"].dropna().astype(str).unique():
        if tag.startswith("AR-"):
            try:
                ar_max = max(ar_max, int(tag.split("-")[1]))
            except (ValueError, IndexError):
                pass
        elif tag.startswith("AI-"):
            try:
                ai_max = max(ai_max, int(tag.split("-")[1]))
            except (ValueError, IndexError):
                pass
    return ar_max + 1, ai_max + 1


# ---------------------------------------------------------------------------
# 5. INCREMENTAL MATCHING ENGINE
# ---------------------------------------------------------------------------
def match_transactions(df):
    """
    Incremental matching:
      - Preserves all rows that already have a MatchGroup
      - Re-classifies and re-matches any row without a MatchGroup
        (this includes new pasted rows AND previously-unmatched rows)
      - Continues match IDs from the highest existing values

    Returns the DataFrame and a stats dict.
    """
    df = df.copy()

    # Make sure all matching columns exist
    for col in MATCHING_COLS:
        if col not in df.columns:
            df[col] = ""

    # Normalize blanks to empty strings for the text columns
    for col in ["MatchGroup", "MatchStatus", "MatchedWith"]:
        df[col] = df[col].fillna("").astype(str).replace("nan", "")

    # Identify which rows need to be (re)processed
    needs_mask = df.apply(needs_matching, axis=1)
    new_count = needs_mask.sum()
    preserved_count = (~needs_mask).sum()

    # Always re-classify rows that need matching (new rows have no RowType)
    rowtype_blank = df["RowType"].apply(is_blank)
    classify_mask = needs_mask | rowtype_blank
    if classify_mask.any():
        df.loc[classify_mask, "RowType"] = df.loc[classify_mask].apply(
            classify_row, axis=1
        )

    # Get next match IDs based on what's already in the file
    ar_id, ai_id = get_next_match_ids(df)
    starting_ar = ar_id
    starting_ai = ai_id

    # Build candidate pools
    available_accruals = df[needs_mask & (df["RowType"] == "Accrual")].copy()
    available_reversals = df[needs_mask & (df["RowType"] == "Reversal")].copy()
    available_invoices = df[needs_mask & (df["RowType"] == "Invoice")].copy()

    available_reversals["AccrualRef"] = available_reversals["GL Journal"].apply(
        extract_accrual_ref
    )

    used_accruals = set()
    used_reversals = set()
    used_invoices = set()

    # --- Pass 1: Accrual <-> Reversal ---
    new_ar_pairs = 0
    for rev_idx, rev in available_reversals.iterrows():
        ref = rev["AccrualRef"]
        if ref is None:
            continue
        candidates = available_accruals[
            (available_accruals["GL Journal"] == ref)
            & (available_accruals["Mapped Vendor"] == rev["Mapped Vendor"])
            & (available_accruals["Company"] == rev["Company"])
            & (abs(available_accruals["Amount"] + rev["Amount"]) < 0.01)
            & (~available_accruals.index.isin(used_accruals))
        ]
        if len(candidates) == 0:
            continue

        acc_idx = candidates.index[0]
        tag = f"AR-{ar_id:04d}"
        ar_id += 1
        new_ar_pairs += 1
        net = round(df.loc[acc_idx, "Amount"] + rev["Amount"], 2)
        status = "Matched - Nets to Zero" if abs(net) < 0.01 else "Matched - Residual"

        used_accruals.add(acc_idx)
        used_reversals.add(rev_idx)

        for idx, mw in [(acc_idx, "Reversal"), (rev_idx, "Accrual")]:
            df.loc[idx, "MatchGroup"] = tag
            df.loc[idx, "MatchedWith"] = mw
            df.loc[idx, "MatchStatus"] = status
            df.loc[idx, "NetCheck"] = net

    # --- Pass 2: Remaining Accrual <-> Invoice (exact amount) ---
    new_ai_pairs = 0
    still_open_acc = available_accruals[~available_accruals.index.isin(used_accruals)]
    for acc_idx, acc in still_open_acc.iterrows():
        candidates = available_invoices[
            (available_invoices["Mapped Vendor"] == acc["Mapped Vendor"])
            & (available_invoices["Company"] == acc["Company"])
            & (abs(available_invoices["Amount"] - acc["Amount"]) < 0.01)
            & (~available_invoices.index.isin(used_invoices))
        ]
        if len(candidates) == 0:
            continue

        inv_idx = candidates.index[0]
        tag = f"AI-{ai_id:04d}"
        ai_id += 1
        new_ai_pairs += 1
        used_invoices.add(inv_idx)
        used_accruals.add(acc_idx)

        df.loc[acc_idx, "MatchGroup"] = tag
        df.loc[acc_idx, "MatchedWith"] = "Invoice"
        df.loc[acc_idx, "MatchStatus"] = "Matched to Invoice (same amount)"
        df.loc[acc_idx, "NetCheck"] = 0.0

        df.loc[inv_idx, "MatchGroup"] = tag
        df.loc[inv_idx, "MatchedWith"] = "Accrual"
        df.loc[inv_idx, "MatchStatus"] = "Matched to Accrual (same amount)"
        df.loc[inv_idx, "NetCheck"] = 0.0

    # --- Label any rows that still need a status ---
    still_needs = df.apply(needs_matching, axis=1)
    for idx in df[still_needs].index:
        rt = df.loc[idx, "RowType"]
        label_map = {
            "Accrual": "UNMATCHED Accrual",
            "Reversal": "UNMATCHED Reversal",
            "Invoice": "Invoice (no accrual match)",
        }
        df.loc[idx, "MatchStatus"] = label_map.get(rt, f"{rt} (not in matching scope)")

    stats = {
        "total_rows": len(df),
        "preserved_rows": preserved_count,
        "processed_rows": new_count,
        "new_ar_pairs": new_ar_pairs,
        "new_ai_pairs": new_ai_pairs,
        "starting_ar_id": starting_ar,
        "starting_ai_id": starting_ai,
        "ending_ar_id": ar_id - 1,
        "ending_ai_id": ai_id - 1,
    }
    return df, stats


# ---------------------------------------------------------------------------
# 6. EXCEL OUTPUT
# ---------------------------------------------------------------------------
# Preferred output order — columns not present in the data are silently skipped
OUTPUT_COLS = [
    "MatchGroup", "MatchStatus", "MatchedWith", "NetCheck", "RowType",
    "GL Source", "Period", "GL Journal", "Acctg Codes", "GL Description",
    "Acct Date", "Vendor Name", "Mapped Vendor", "Invoice #", "Inv Date",
    "Cust #", "AR Location", "AR Trans Type",
    "Company", "Company Name", "Account", "Acct Name",
    "PeriodDate", "Year", "Month Name", "PeriodLabel",
    "Category", "Category Filter", "Reference Detail",
    "ENTERED_DR", "ENTERED_CR", "TOTAL", "Amount",
]

COL_WIDTHS = {
    "MatchGroup": 12, "MatchStatus": 30, "MatchedWith": 14, "NetCheck": 12,
    "RowType": 11, "GL Source": 12, "Period": 10, "GL Journal": 55,
    "Acctg Codes": 16, "GL Description": 30, "Acct Date": 12,
    "Vendor Name": 30, "Mapped Vendor": 30, "Invoice #": 16,
    "Inv Date": 12, "Cust #": 10, "AR Location": 14, "AR Trans Type": 14,
    "PeriodDate": 12, "PeriodLabel": 10, "Category": 28,
    "Amount": 14, "ENTERED_DR": 13, "ENTERED_CR": 13, "TOTAL": 13,
    "Company Name": 24, "Reference Detail": 40,
}

MONEY_FMT = '#,##0.00;(#,##0.00);"-"'

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
BLUE = PatternFill("solid", fgColor="BDD7EE")
LBLUE = PatternFill("solid", fgColor="D9E2F3")
GRAY = PatternFill("solid", fgColor="F2F2F2")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=9)
THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))

MONEY_COLUMNS = {"ENTERED_DR", "ENTERED_CR", "TOTAL", "Amount", "NetCheck"}


def status_to_fill(status):
    s = str(status)
    if "Nets to Zero" in s:
        return GREEN
    if "UNMATCHED" in s:
        return RED
    if "Matched to" in s:
        return LBLUE
    if "no accrual match" in s:
        return BLUE
    if "not in matching scope" in s:
        return GRAY
    return None


def write_excel(df, stats, outpath, sheet_name):
    """Write the matched DataFrame to a formatted Excel file."""
    cols = [c for c in OUTPUT_COLS if c in df.columns]
    # Also include any original columns not in OUTPUT_COLS so data isn't lost
    extra = [c for c in df.columns if c not in cols and not c.startswith("_")]
    cols = cols + extra

    df_out = df[cols]

    # Sort: matched pairs grouped together, then unmatched at bottom
    df_out = df_out.copy()
    df_out["_s1"] = df_out["MatchGroup"].apply(lambda x: 0 if x and str(x).strip() else 1)
    df_out["_s2"] = df_out["MatchGroup"]
    df_out["_s3"] = df_out["RowType"].map({"Accrual": 0, "Reversal": 1, "Invoice": 2}).fillna(3)

    sort_cols = ["_s1", "_s2", "_s3"]
    if "Period" in df_out.columns:
        sort_cols.append("Period")
    elif "PeriodDate" in df_out.columns:
        sort_cols.append("PeriodDate")

    df_out = df_out.sort_values(sort_cols).drop(columns=["_s1", "_s2", "_s3"])

    df_out.to_excel(outpath, index=False, sheet_name=sheet_name)

    wb = load_workbook(outpath)
    ws = wb[sheet_name]

    # Header row
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Build column index lookups from the actual written columns
    written_cols = list(df_out.columns)
    status_ci = written_cols.index("MatchStatus") + 1 if "MatchStatus" in written_cols else None
    money_cis = [written_cols.index(c) + 1 for c in MONEY_COLUMNS if c in written_cols]

    for row in range(2, ws.max_row + 1):
        fill = status_to_fill(ws.cell(row=row, column=status_ci).value) if status_ci else None
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if col in money_cis:
                cell.number_format = MONEY_FMT

    for i, cn in enumerate(written_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(cn, 14)

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(outpath)


# ---------------------------------------------------------------------------
# 7. SHEET DETECTION
# ---------------------------------------------------------------------------
def find_sheet(input_path, requested_sheet=None):
    """Find the right sheet to read - either user-specified or by trying defaults."""
    xls = pd.ExcelFile(input_path)
    available = xls.sheet_names

    if requested_sheet:
        if requested_sheet in available:
            return requested_sheet
        print(f"Error: sheet '{requested_sheet}' not found in workbook")
        print(f"Available sheets: {available}")
        sys.exit(1)

    for default in DEFAULT_SHEET_NAMES:
        if default in available:
            return default

    return available[0]


# ---------------------------------------------------------------------------
# 8. MAIN
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.isfile(input_path):
        print(f"Error: file not found: {input_path}")
        sys.exit(1)

    requested_sheet = sys.argv[2] if len(sys.argv) >= 3 and sys.argv[2] else None
    vendor_csv = sys.argv[3] if len(sys.argv) >= 4 else DEFAULT_VENDOR_CSV
    sheet_name = find_sheet(input_path, requested_sheet)

    # Build timestamped output path
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = os.path.dirname(input_path) or "."
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    output_path = os.path.join(base_dir, f"{base_name}_{timestamp}.xlsx")

    print(f"Input:  {input_path}")
    print(f"Sheet:  {sheet_name}")
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    print(f"  {len(df)} rows loaded")

    # Normalise column names to what the matching engine expects
    df = normalise_columns(df)
    print(f"  Columns after normalisation: {list(df.columns)}")

    # Extract vendors from GL Description using the external vendor list
    vendor_list = load_vendor_list(vendor_csv)
    if vendor_list:
        print(f"  Loaded {len(vendor_list)} vendors from {vendor_csv}")
        df = apply_vendor_extraction(df, vendor_list)

    print("Matching...")
    df_matched, stats = match_transactions(df)

    print("Writing output...")
    write_excel(df_matched, stats, output_path, sheet_name)

    print(f"\nDone -> {output_path}")
    print(f"  Total rows:           {stats['total_rows']:>6,}")
    print(f"  Preserved (existing): {stats['preserved_rows']:>6,}")
    print(f"  Processed this run:   {stats['processed_rows']:>6,}")
    print(f"  New AR pairs:         {stats['new_ar_pairs']:>6,}", end="")
    if stats["new_ar_pairs"]:
        print(f"  (AR-{stats['starting_ar_id']:04d} through AR-{stats['ending_ar_id']:04d})")
    else:
        print()
    print(f"  New AI pairs:         {stats['new_ai_pairs']:>6,}", end="")
    if stats["new_ai_pairs"]:
        print(f"  (AI-{stats['starting_ai_id']:04d} through AI-{stats['ending_ai_id']:04d})")
    else:
        print()


if __name__ == "__main__":
    main()
